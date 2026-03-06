import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

def gerar_template_blindado():
    nome_arquivo = "Planilha Robo.xlsx"
    
    # Se já existir um corrompido com esse nome, ele apaga antes de criar o novo
    if os.path.exists(nome_arquivo):
        try:
            os.remove(nome_arquivo)
        except:
            print(f"Erro: Feche o arquivo {nome_arquivo} antes de gerar um novo!")
            return

    print("Criando nova planilha blindada...")
    wb = openpyxl.Workbook()
    
    # 1. CRIANDO A ABA "ROBO" (Principal)
    ws_robo = wb.active
    ws_robo.title = "ROBO"
    
    # Estilos visuais (Azul Escuro com letra Branca e Negrito para o Cabeçalho)
    fundo_azul = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)
    centralizado = Alignment(horizontal="center", vertical="center")
    
    # Configurações do E-mail (Linhas 4, 5 e 6)
    ws_robo["B4"] = "Assunto:"
    ws_robo["B4"].font = Font(bold=True)
    ws_robo["C4"] = "Atualização - Cartão Alelo"
    
    ws_robo["B5"] = "CC:"
    ws_robo["B5"].font = Font(bold=True)
    ws_robo["C5"] = "" # Deixe em branco pro RH preencher se quiser
    
    ws_robo["B6"] = "BCC:"
    ws_robo["B6"].font = Font(bold=True)
    ws_robo["C6"] = "" # Deixe em branco pro RH preencher se quiser

    # Cabeçalhos na Linha 8
    cabecalhos = ['ID', 'Matricula', 'Nome', 'Cargo', 'Código de Rastreio', 'Data de Postagem', 'Status', 'Obs', 'Email', 'Enviar']
    
    for col_idx, texto in enumerate(cabecalhos, start=1):
        celula = ws_robo.cell(row=8, column=col_idx, value=texto)
        celula.fill = fundo_azul
        celula.font = fonte_branca
        celula.alignment = centralizado
        
    # 2. CRIANDO A ABA "PROCX"
    ws_procx = wb.create_sheet("PROCX")
    
    # Cabeçalhos do PROCX na Linha 1
    for col_idx, texto in enumerate(cabecalhos, start=1):
        celula = ws_procx.cell(row=1, column=col_idx, value=texto)
        celula.fill = fundo_azul
        celula.font = fonte_branca
        celula.alignment = centralizado

    # INJETANDO AS FÓRMULAS DO PROCX NA LINHA 2!
    f_matricula = "=PROCX(C2; '[05032026_Relação de Ativos_ECP.XLSX]ATIVOS'!$D:$D; '[05032026_Relação de Ativos_ECP.XLSX]ATIVOS'!$C:$C; \"Não encontrado\")"
    f_cargo = "=ARRUMAR(PROCX(C2; '[05032026_Relação de Ativos_ECP.XLSX]ATIVOS'!$D:$D; '[05032026_Relação de Ativos_ECP.XLSX]ATIVOS'!$P:$P; \"Não encontrado\"))"
    f_email = "=SUBSTITUIR(MINUSCULA(ARRUMAR(PROCX(C2; '[e mail 19.02.XLSX]Sheet1'!$B:$B; '[e mail 19.02.XLSX]Sheet1'!$C:$C; \"\"))); \"viavarejo.com.br\"; \"casasbahia.com.br\")"
    
    ws_procx["B2"] = f_matricula
    ws_procx["D2"] = f_cargo
    ws_procx["I2"] = f_email
    
    # Deixando o visual bonito (largura das colunas)
    larguras = {"A": 5, "B": 15, "C": 40, "D": 30, "E": 20, "F": 18, "G": 15, "H": 25, "I": 35, "J": 8}
    for aba in [ws_robo, ws_procx]:
        for letra, largura in larguras.items():
            aba.column_dimensions[letra].width = largura

    # Salvando o troféu
    wb.save(nome_arquivo)
    print(f"✅ {nome_arquivo} criado com sucesso! As fórmulas do PROCX já estão na linha 2 da aba PROCX.")

if __name__ == "__main__":
    gerar_template_blindado()