import os
import sys
import json
import threading
import time
import re
import shutil
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Querybox
from playwright.sync_api import sync_playwright
from PIL import Image, ImageTk

# --- A TRAVA DE PORTABILIDADE (Deixa o .exe rodar em qualquer PC) ---
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "0"

# --- FUNÇÃO MÁGICA PARA A LOGO NO .EXE ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- CONFIGURAÇÕES DE DIRETÓRIO DINÂMICAS ---
if getattr(sys, 'frozen', False):
    PASTA_BASE = os.path.dirname(sys.executable)
else:
    PASTA_BASE = os.path.dirname(os.path.abspath(__file__))

CAMINHO_PERFIL_CHROME = os.path.join(PASTA_BASE, "perfil_chrome")
ARQUIVO_CONFIG = os.path.join(PASTA_BASE, "config_teams.json")

ARQUIVOS_TEMPLATES = {
    "sedex": os.path.join(PASTA_BASE, "template_sedex.txt"),
    "scs": os.path.join(PASTA_BASE, "template_scs.txt"),
    "berrini": os.path.join(PASTA_BASE, "template_berrini.txt"),
    "reaviso_scs": os.path.join(PASTA_BASE, "template_reaviso_scs.txt"),
    "reaviso_berrini": os.path.join(PASTA_BASE, "template_reaviso_berrini.txt"),
    "personalizada": os.path.join(PASTA_BASE, "template_personalizada.txt") 
}

TEXTOS_PADROES = {
    "sedex": """Boa tarde.
Olá {primeiro_nome}, tudo bem?

É um prazer tê-lo em nossa CIA como um de nossos (as) colaboradores (as), nós do time de benefícios temos uma excelente noticia para você!
Informamos que o seu cartão *Alelo*, foi entregue em nosso HUB SCS e iremos redireciona-lo via SEDEX ao endereço cadastrado em sistema.

Segue abaixo código de rastreio, para acessar, basta entrar no link: https://www.correios.com.br/

[TABELA_RASTREIO]

Obs.: Esta é uma mensagem automática. Por favor não responda este e-mail.

Atenciosamente,
*Grupo Casas Bahia - Gente, Gestão e Sustentabilidade*
Operações de Benefícios""",

    "scs": """Olá {primeiro_nome}, tudo bem?
    
Informamos que o seu *Cartão Alelo* foi recebido em nosso Hub e está disponível para a retirada.

📌 *Atenção colaboradores alocados sistemicamente no Hub SCS – Filial 1580, segue abaixo informações de retirada.*

📍 *Local de retirada:*
*Hub SCS - 2° andar – Mesa n° 636 (Terceira baia a frente das salas de reuniões.)*

🕒 *Horários de atendimento:*
SEG | QUA | QUI | SEX (Sem entregas as terças-feiras)
*Manhã:* das 9h às 11h30
*Tarde:* das 12h30 às 16h

Pedimos atenção à data para garantir o recebimento do cartão no local indicado.

❓ *Dúvidas frequentes:*

*• Onde identificar minha Filial de cadastro?*
Para consulta sua filial de cadastro, acesse o Portal do Colaborador> Meu perfil> a informação estará abaixo do seu nome, departamento e diretoria.

*• Posso solicitar que outra pessoa retire meu cartão?*
Sim. A retirada pode ser feita por outro colaborador, desde que informe seu *nome completo e matrícula* no momento da retirada.

Atenciosamente;
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    "berrini": """Olá {primeiro_nome}, tudo bem?
    
Informamos que o seu *Cartão Alelo* foi recebido em nosso Hub e encontra-se disponível para retirada conforme as orientações abaixo.

📌 *Colaboradores alocados sistemicamente na Estação Casas Bahia (Berrini) – Filiais 01 | 650 | 1968:*
A retirada deve ocorrer no dia *06/03/2026*, de forma pontual, conforme a ida de um de nossos portadores ao local.

📍 *Local de retirada:*
*Hub Estação Casas Bahia (Berrini) - 4º andar - sala de Bem-Estar (ao lado do Espaço Viver Bem)*

🕒 *Horário de atendimento:*
9h30 às 17h

Pedimos atenção a data informada referente ao plantão de entrega.

❓ *Dúvidas frequentes:*

*• Onde identificar minha Filial de cadastro?*
Para consulta sua filial de cadastro, acesse o Portal do Colaborador> Meu perfil> a informação estará abaixo do seu nome, departamento e diretoria.

*• Posso solicitar que outra pessoa retire meu cartão?*
Sim. A retirada pode ser feita por outro colaborador, desde que informe seu *nome completo e matrícula* no momento da retirada.
Caso já tenha retirado o seu cartão Alelo, por favor desconsiderar este e-mail!!

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    "reaviso_scs": """Olá {primeiro_nome}, tudo bem?
 
Identificamos que você ainda não realizou a retirada do seu Cartão Alelo.

Gostaríamos de lembrá-lo(a) que o seu cartão está disponível para retirada hoje, 06/03/2026.
 
📍 *Local de retirada:*
Hub SCS – 2º andar – Mesa 636.
 
🕒 *Horário de atendimento:*
09h às 15h.

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    "reaviso_berrini": """Olá {primeiro_nome}, tudo bem?
 
Identificamos que você ainda não realizou a retirada do seu Cartão Alelo.

Gostaríamos de lembrá-lo(a) que o seu cartão está disponível para retirada hoje, 06/03/2026.
 
📍 *Local de retirada:*
Hub Estação Casas Bahia (Berrini) - 4º andar - sala de Bem-Estar (ao lado do Espaço Viver Bem)
 
🕒 *Horário de atendimento:*
09h às 15h.

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    "personalizada": """Olá {primeiro_nome}, tudo bem?

[Apague este texto e digite o seu comunicado aqui. Você pode usar a tag {primeiro_nome} quantas vezes quiser no texto para o robô trocar pelo nome da pessoa.]

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br"""
}

def carregar_usuario_salvo():
    if os.path.exists(ARQUIVO_CONFIG):
        try:
            with open(ARQUIVO_CONFIG, "r", encoding="utf-8") as f:
                return json.load(f).get("usuario", "Desconectado")
        except Exception: return "Desconectado"
    return "Desconectado"

def salvar_usuario(nome):
    with open(ARQUIVO_CONFIG, "w", encoding="utf-8") as f:
        json.dump({"usuario": nome}, f)

def carregar_template(tipo):
    caminho = ARQUIVOS_TEMPLATES[tipo]
    if os.path.exists(caminho):
        try:
            with open(caminho, "r", encoding="utf-8") as f: return f.read()
        except Exception: return TEXTOS_PADROES[tipo]
    else:
        with open(caminho, "w", encoding="utf-8") as f: f.write(TEXTOS_PADROES[tipo])
        return TEXTOS_PADROES[tipo]

def salvar_template(tipo, texto):
    with open(ARQUIVOS_TEMPLATES[tipo], "w", encoding="utf-8") as f: f.write(texto)

def limpar_trava_navegador(caminho_perfil):
    lock_file = os.path.join(caminho_perfil, "SingletonLock")
    if os.path.exists(lock_file):
        try: os.remove(lock_file)
        except Exception: pass

def robo_login(app_gui, nome_usuario):
    app_gui.atualizar_status("🌐 Abrindo Microsoft Teams para login...", INFO)
    app_gui.btn_login.config(state="disabled")
    limpar_trava_navegador(CAMINHO_PERFIL_CHROME)
    try:
        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(user_data_dir=CAMINHO_PERFIL_CHROME, channel="chrome", headless=False, args=["--start-maximized"])
            page = context.pages[0] if context.pages else context.new_page()
            page.goto("https://teams.microsoft.com/v2/")
            app_gui.atualizar_status("⏳ Faça o login no Chrome. Feche a janela quando o Teams carregar.", WARNING)
            try: page.wait_for_event("close", timeout=0) 
            except: pass
            context.close()
            salvar_usuario(nome_usuario)
            app_gui.var_usuario.set(f"👤 Logado como: {nome_usuario}")
            app_gui.atualizar_status("✅ Sessão salva com sucesso!", SUCCESS)
            app_gui.btn_login.config(state="normal")
    except Exception:
        app_gui.atualizar_status("Login cancelado ou navegador fechado.", INFO)
        app_gui.btn_login.config(state="normal")

def robo_disparos(app_gui, caminho_arquivo, template_texto_puro):
    app_gui.atualizar_status("Iniciando Robô do Teams...", INFO)
    app_gui.btn_iniciar.config(state="disabled")
    
    try:
        wb = load_workbook(caminho_arquivo)
        sh_ROBO = wb["ROBO"]
        df_completo = pd.read_excel(caminho_arquivo, sheet_name="ROBO", header=7)
        df_completo.columns = df_completo.columns.str.strip()
        df_pendentes = df_completo[(df_completo['Status'] != 'Enviado') & (df_completo['Enviar'] == 'x')]
        
        if df_pendentes.empty:
            app_gui.atualizar_status("Nenhuma mensagem pendente na planilha!", WARNING)
            app_gui.btn_iniciar.config(state="normal")
            return

        limpar_trava_navegador(CAMINHO_PERFIL_CHROME)

        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=CAMINHO_PERFIL_CHROME, channel="chrome", headless=False,
                args=["--start-maximized", "--disable-popup-blocking"],
                permissions=["clipboard-read", "clipboard-write"], viewport=None
            )
            page = context.pages[0] if context.pages else context.new_page()
            app_gui.atualizar_status("Carregando o Teams principal...", INFO)
            page.goto("https://teams.microsoft.com/v2/")
            page.wait_for_load_state("domcontentloaded")
            time.sleep(5) 
            page.reload()
            page.wait_for_load_state("domcontentloaded")
            time.sleep(5) 
            
            total = len(df_pendentes)
            enviados = 0
            lista_falhas = []
            
            for index, row in df_pendentes.iterrows():
                email_destino = str(row.get('Email', '')).strip()
                if not email_destino or email_destino.lower() == 'nan':
                    continue 
                    
                nome = str(row.get('Nome', 'Colaborador')).strip()
                rastreio = str(row.get('Código de Rastreio', 'N/D'))
                matricula = str(row.get('Matricula', 'N/D')).split('.')[0] 
                cargo = str(row.get('Cargo', 'N/D'))
                
                try:
                    data_postagem = pd.to_datetime(row.get('Data de Postagem')).strftime('%d/%m/%Y')
                except:
                    data_postagem = str(row.get('Data de Postagem', 'N/D')).split(' ')[0]
                    
                primeiro_nome = nome.split()[0].capitalize() if nome and nome.lower() != 'nan' else "Colaborador"
                prefixo_email = email_destino.split('@')[0]
                
                texto_preparado = template_texto_puro.replace("{primeiro_nome}", primeiro_nome)
                html_final = texto_preparado.replace('\n', '<br>')
                html_final = re.sub(r'\*(.*?)\*', r'<b>\1</b>', html_final)
                
                if "[TABELA_RASTREIO]" in html_final:
                    tabela_html = f"""<br><table border="1" style="border-collapse: collapse; width: 100%; text-align: center; font-family: sans-serif; font-size: 12px;">
                      <tr style="background-color: #9bc2e6; font-weight: bold;">
                        <th style="padding: 6px; border: 1px solid black;">Matricula</th>
                        <th style="padding: 6px; border: 1px solid black;">Nome</th>
                        <th style="padding: 6px; border: 1px solid black;">Cargo</th>
                        <th style="padding: 6px; border: 1px solid black;">Código de Rastreio</th>
                        <th style="padding: 6px; border: 1px solid black;">Data de Postagem</th>
                      </tr>
                      <tr>
                        <td style="padding: 6px; border: 1px solid black;">{matricula}</td>
                        <td style="padding: 6px; border: 1px solid black;">{nome}</td>
                        <td style="padding: 6px; border: 1px solid black;">{cargo}</td>
                        <td style="padding: 6px; border: 1px solid black;">{rastreio}</td>
                        <td style="padding: 6px; border: 1px solid black;">{data_postagem}</td>
                      </tr>
                    </table><br>"""
                    html_final = html_final.replace("[TABELA_RASTREIO]", tabela_html)
                
                app_gui.atualizar_status(f"🚀 Enviando para: {primeiro_nome}... ({enviados+1}/{total})", PRIMARY)
                linha_excel = index + 9
                coluna_status = 7
                for col in range(1, sh_ROBO.max_column + 1):
                    if str(sh_ROBO.cell(row=8, column=col).value).strip() == "Status": coluna_status = col; break

                try:
                    page.bring_to_front()
                    
                    barra_pesquisa = page.locator('[data-tid="AUTOSUGGEST_INPUT"]')
                    barra_pesquisa.wait_for(state="visible", timeout=7000) 
                    
                    barra_pesquisa.click(force=True)
                    barra_pesquisa.focus()
                    
                    barra_pesquisa.press("Control+a")
                    barra_pesquisa.press("Backspace")
                    time.sleep(0.5) 
                    
                    achou_pessoa = False

                    # ====================================================================
                    # TENTATIVA 1: BUSCA POR E-MAIL E CLIQUE ESPECÍFICO
                    # ====================================================================
                    barra_pesquisa.type(email_destino, delay=50)
                    time.sleep(3) # Tempo para o Teams carregar a lista do AD
                    
                    try:
                        # Tenta clicar no card que tenha o NOME COMPLETO da pessoa (ignora a pílula pequena)
                        opcao_correta = page.locator('div[role="option"]').filter(has_text=re.compile(re.escape(nome), re.IGNORECASE)).first
                        opcao_correta.wait_for(state="visible", timeout=2000)
                        opcao_correta.click()
                        time.sleep(1.5) 
                        achou_pessoa = True
                    except:
                        try:
                            # Tenta clicar no card que tenha o PREFIXO DO EMAIL (ex: william.bernst)
                            opcao_correta = page.locator('div[role="option"]').filter(has_text=re.compile(re.escape(prefixo_email), re.IGNORECASE)).first
                            opcao_correta.wait_for(state="visible", timeout=2000)
                            opcao_correta.click()
                            time.sleep(1.5) 
                            achou_pessoa = True
                        except:
                            pass # Falhou a busca por E-mail, vai para a Tentativa 2
                            
                    # ====================================================================
                    # TENTATIVA 2: SE FALHOU O E-MAIL, BUSCA PELO NOME COMPLETO
                    # ====================================================================
                    if not achou_pessoa:
                        print(f"Não achou por e-mail. Tentando pelo Nome Completo: {nome}")
                        
                        # Limpa a barra
                        barra_pesquisa.click(force=True)
                        barra_pesquisa.focus()
                        barra_pesquisa.press("Control+a")
                        barra_pesquisa.press("Backspace")
                        time.sleep(0.5)
                        
                        # Digita o Nome Completo
                        barra_pesquisa.type(str(nome), delay=50)
                        time.sleep(4) # Espera maior pro AD buscar o nome
                        
                        try:
                            # Tenta clicar no resultado que tem o nome exato
                            opcao_correta = page.locator('div[role="option"]').filter(has_text=re.compile(re.escape(nome), re.IGNORECASE)).first
                            opcao_correta.wait_for(state="visible", timeout=2000)
                            opcao_correta.click()
                            time.sleep(1.5)
                            achou_pessoa = True
                        except:
                            try:
                                # Força bruta segura: pega a primeira opção que tenha pelo menos o PRIMEIRO NOME
                                primeiro_nome_busca = str(nome).strip().split()[0]
                                opcao_correta = page.locator('div[role="option"]').filter(has_text=re.compile(re.escape(primeiro_nome_busca), re.IGNORECASE)).first
                                opcao_correta.wait_for(state="visible", timeout=2000)
                                opcao_correta.click()
                                time.sleep(1.5)
                                achou_pessoa = True
                            except:
                                pass # Falhou em tudo.

                    # ====================================================================
                    # TRAVA DE SEGURANÇA: ABORTAR SE NÃO ACHOU NINGUÉM
                    # ====================================================================
                    if not achou_pessoa:
                        # O raise Exception pula direto para o bloco 'except' lá embaixo,
                        # impedindo que o código desça e digite a mensagem no chat atual.
                        raise Exception(f"Pessoa '{nome}' não encontrada na busca. Abortando envio.")


                    # Se chegou aqui, é porque ele achou a pessoa e clicou com sucesso!
                    caixa_texto = page.locator('[data-tid="ckeditor"]')
                    caixa_texto.wait_for(state="visible", timeout=7000) 
                    
                    caixa_texto.click(force=True)
                    caixa_texto.focus()
                    time.sleep(0.5)
                    
                    caixa_texto.press("Control+a")
                    caixa_texto.press("Backspace")
                    time.sleep(0.5)
                    
                    page.evaluate("""async (htmlContent) => {
                        const blobHtml = new Blob([htmlContent], { type: 'text/html' });
                        const blobText = new Blob([htmlContent], { type: 'text/plain' });
                        const clipboardItem = new ClipboardItem({'text/html': blobHtml, 'text/plain': blobText});
                        await navigator.clipboard.write([clipboardItem]);
                    }""", html_final)
                    
                    time.sleep(1) 
                    
                    caixa_texto.focus() 
                    page.keyboard.press("Control+v")
                    time.sleep(3) 
                    
                    try:
                        btn_enviar = page.locator('[data-tid="send-message-button"], button[aria-label="Enviar"], button[aria-label="Send"]').first
                        btn_enviar.wait_for(state="visible", timeout=3000)
                        btn_enviar.click(force=True)
                        time.sleep(2)
                    except:
                        caixa_texto.focus()
                        time.sleep(0.5)
                        page.keyboard.press("End") 
                        time.sleep(0.5)
                        page.keyboard.type(" ") 
                        time.sleep(0.5)
                        page.keyboard.press("Enter") 
                        time.sleep(2) 
                            
                    sh_ROBO.cell(row=linha_excel, column=coluna_status).value = "Enviado"
                    wb.save(caminho_arquivo) 
                    enviados += 1
                    
                except Exception as e:
                    print(f"Falha ao enviar para {nome}: {e}")
                    sh_ROBO.cell(row=linha_excel, column=coluna_status).value = "Não Encontrado"
                    wb.save(caminho_arquivo)
                    lista_falhas.append(str(nome))
                    
                    # Limpa a sujeira da barra de pesquisa e fecha o popup pra não travar o próximo
                    try:
                        barra_pesquisa.click(force=True)
                        barra_pesquisa.press("Control+a")
                        barra_pesquisa.press("Backspace")
                    except:
                        pass
                    
                    page.keyboard.press("Escape")
                    page.locator("body").click(force=True)
                    time.sleep(2)

            context.close()
            app_gui.btn_iniciar.config(state="normal")
            
            if lista_falhas:
                app_gui.atualizar_status(f"⚠️ Concluído com {len(lista_falhas)} falha(s).", WARNING)
                nomes_falha = "\n".join(lista_falhas)
                messagebox.showwarning("Atenção - Relatório de Disparos", f"Processo finalizado!\n\n✅ Sucesso: {enviados}\n❌ Falhas: {len(lista_falhas)}\n\nAs seguintes pessoas não foram encontradas no Teams:\n\n{nomes_falha}\n\nO status na planilha foi alterado para 'Não Encontrado'.")
            else:
                app_gui.atualizar_status(f"🎉 Concluído com sucesso!", SUCCESS)
                messagebox.showinfo("Sucesso Total", "Todos os disparos foram concluídos com sucesso!")

    except Exception as e:
        print(f"Erro Fatal: {e}")
        app_gui.atualizar_status(f"Erro Crítico: {str(e)[:50]}...", DANGER)
        app_gui.btn_iniciar.config(state="normal")

class AppTeams(tb.Window):
    def __init__(self):
        super().__init__(themename="yeti") 
        self.title("Robô Envio Teams Alelo") 
        self.geometry("1200x850") 
        self.resizable(False, False)
        
        caminho_icone = resource_path("logo.png")
        if os.path.exists(caminho_icone):
            try:
                img_icon = Image.open(caminho_icone)
                icone = ImageTk.PhotoImage(img_icon)
                self.iconphoto(False, icone)
            except Exception:
                pass 
        
        self.caminho_planilha = ""
        self.tipo_ativo = "sedex" 
        
        usuario_salvo = carregar_usuario_salvo()
        texto_usuario = f"👤 Logado como: {usuario_salvo}" if usuario_salvo != "Desconectado" else "🔴 Desconectado"
        self.var_usuario = tb.StringVar(value=texto_usuario)
        self.var_status = tb.StringVar(value="Pronto para iniciar.")
        self.var_tipo_msg = tb.StringVar(value="sedex")
        
        self.construir_interface()

    def construir_interface(self):
        frame_main = tb.Frame(self, padding=20)
        frame_main.pack(fill=BOTH, expand=True)
        
        frame_header = tb.Frame(frame_main)
        frame_header.pack(fill=X, pady=(0, 15))
        
        frame_titles = tb.Frame(frame_header)
        frame_titles.pack(side=LEFT)
        
        lbl_titulo = tb.Label(frame_titles, text="Automação Microsoft Teams", font=("Segoe UI", 20, "bold"), bootstyle=PRIMARY)
        lbl_titulo.pack(anchor=W)
        lbl_subtitulo = tb.Label(frame_titles, text="Módulo de Disparos - Cartão Alelo", font=("Segoe UI", 10), foreground="gray")
        lbl_subtitulo.pack(anchor=W)
        
        btn_ajuda = tb.Button(frame_header, text="Como utilizar❓", bootstyle="primary", command=self.mostrar_ajuda)
        btn_ajuda.pack(side=RIGHT, anchor=N)
        
        frame_rodape = tb.Frame(frame_main)
        frame_rodape.pack(fill=X, side=BOTTOM)
        self.lbl_status = tb.Label(frame_rodape, textvariable=self.var_status, font=("Segoe UI", 10, "bold"), bootstyle=SECONDARY)
        self.lbl_status.pack(side=LEFT)

        self.btn_iniciar = tb.Button(frame_main, text="▶ INICIAR DISPAROS", bootstyle="success", padding=10, command=self.iniciar_disparos)
        self.btn_iniciar.pack(fill=X, side=BOTTOM, pady=(15, 15))

        frame_content = tb.Frame(frame_main)
        frame_content.pack(fill=BOTH, expand=True)

        frame_left = tb.Frame(frame_content)
        frame_left.pack(side=LEFT, fill=Y, expand=False, padx=(0, 15))

        frame_login = tb.LabelFrame(frame_left, text=" 1. Conta e Autenticação ")
        frame_login.pack(fill=X, pady=(0, 15), ipadx=5, ipady=5)
        lbl_logado = tb.Label(frame_login, textvariable=self.var_usuario, font=("Segoe UI", 11, "bold"))
        lbl_logado.pack(anchor=W, padx=15, pady=(10, 5))
        self.btn_login = tb.Button(frame_login, text="🔌 Conectar Nova Conta", bootstyle="primary", command=self.iniciar_processo_login)
        self.btn_login.pack(anchor=W, padx=15, pady=(0, 10))

        frame_arquivo = tb.LabelFrame(frame_left, text=" 2. Base de Disparos ")
        frame_arquivo.pack(fill=X, pady=(0, 15), ipadx=5, ipady=5)
        
        frame_botoes_arquivo = tb.Frame(frame_arquivo)
        frame_botoes_arquivo.pack(anchor=W, fill=X, padx=15, pady=(10, 5))
        
        self.btn_procurar = tb.Button(frame_botoes_arquivo, text="📂 Escolher Planilha", bootstyle="primary", command=self.selecionar_arquivo)
        self.btn_procurar.pack(side=LEFT, padx=(0, 10))
        
        self.btn_limpar = tb.Button(frame_botoes_arquivo, text="🧹 Limpar Status", bootstyle="warning", command=self.limpar_status_planilha)
        self.btn_limpar.pack(side=LEFT)
        
        self.lbl_caminho = tb.Label(frame_arquivo, text="Nenhuma planilha (.xlsx) selecionada", font=("Segoe UI", 9, "italic"), foreground="black", wraplength=280)
        self.lbl_caminho.pack(anchor=W, fill=X, expand=True, padx=15, pady=(0, 10))

        frame_tipo = tb.LabelFrame(frame_left, text=" 3. Tipo de Comunicação ")
        frame_tipo.pack(fill=X, pady=(0, 0), ipadx=5, ipady=5)
        tb.Radiobutton(frame_tipo, text="Envio Correios (Sedex)", variable=self.var_tipo_msg, value="sedex", command=self.trocar_aba).pack(anchor=W, padx=15, pady=(10, 5))
        tb.Radiobutton(frame_tipo, text="Retirada Presencial (Hub SCS)", variable=self.var_tipo_msg, value="scs", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Radiobutton(frame_tipo, text="Retirada Presencial (Hub Berrini)", variable=self.var_tipo_msg, value="berrini", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Separator(frame_tipo).pack(fill=X, padx=15, pady=5)
        tb.Radiobutton(frame_tipo, text="Re-aviso (Hub SCS)", variable=self.var_tipo_msg, value="reaviso_scs", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Radiobutton(frame_tipo, text="Re-aviso (Hub Berrini)", variable=self.var_tipo_msg, value="reaviso_berrini", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Separator(frame_tipo).pack(fill=X, padx=15, pady=5)
        tb.Radiobutton(frame_tipo, text="Mensagem Personalizada (Livre)", variable=self.var_tipo_msg, value="personalizada", command=self.trocar_aba).pack(anchor=W, padx=15, pady=(5, 10))

        frame_right = tb.Frame(frame_content)
        frame_right.pack(side=LEFT, fill=BOTH, expand=True)

        frame_msg = tb.LabelFrame(frame_right, text=" 4. Pré-visualização e Edição (Texto Simples) ")
        frame_msg.pack(fill=BOTH, expand=True, ipadx=5, ipady=5)
        
        lbl_dica = tb.Label(frame_msg, text="💡 DICA: Escreva normalmente. Para negrito, use *asteriscos*. O robô ajusta as linhas automaticamente.", font=("Segoe UI", 9, "bold"), bootstyle=PRIMARY)
        lbl_dica.pack(anchor=W, padx=15, pady=(5,0))
        
        scroll_txt = tb.Scrollbar(frame_msg)
        scroll_txt.pack(side=RIGHT, fill=Y, pady=5, padx=(0,10))
        self.txt_mensagem = tb.Text(frame_msg, font=("Segoe UI", 10), wrap="word", yscrollcommand=scroll_txt.set)
        self.txt_mensagem.pack(side=LEFT, fill=BOTH, expand=True, padx=(15,0), pady=5)
        scroll_txt.config(command=self.txt_mensagem.yview)
        
        self.txt_mensagem.insert("1.0", carregar_template("sedex"))

    def mostrar_ajuda(self):
        texto_ajuda = (
            "PASSO A PASSO DE USO:\n\n"
            "1. AUTENTICAÇÃO:\n"
            "Clique em 'Conectar Nova Conta'. Faça o login normalmente no navegador e feche a janela assim que o Teams carregar.\n\n"
            "2. PLANILHA DE DADOS:\n"
            "Clique em 'Escolher Planilha' e selecione a sua base. A planilha precisa ter a coluna 'Enviar' marcada com 'x'. (O robô cria uma cópia de segurança automática para não corromper sua base original!).\n\n"
            "3. MENSAGEM:\n"
            "Escolha o tipo de disparo. Você pode editar o texto na tela preta. Se usar a opção 'Mensagem Personalizada', você pode escrever o que quiser e usar a tag {primeiro_nome} para o robô chamar o colaborador pelo nome automaticamente.\n\n"
            "4. EXECUTAR:\n"
            "Clique em 'Iniciar Disparos' e deixe o mouse parado. Dica: Se quiser testar o robô de novo, clique em 'Limpar Status' para apagar os registros da planilha."
        )
        messagebox.showinfo("Guia Rápido - Robô do Teams", texto_ajuda)

    def trocar_aba(self):
        texto_atual = self.txt_mensagem.get("1.0", tk.END).strip()
        salvar_template(self.tipo_ativo, texto_atual)
        novo_tipo = self.var_tipo_msg.get()
        self.tipo_ativo = novo_tipo
        self.txt_mensagem.delete("1.0", tk.END)
        self.txt_mensagem.insert("1.0", carregar_template(novo_tipo))

    def atualizar_status(self, texto, estilo=INFO):
        self.var_status.set(texto)
        self.lbl_status.config(bootstyle=estilo)
        self.update_idletasks()

    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(title="Selecione a Planilha", filetypes=[("Planilhas do Excel", "*.xlsx")])
        if arquivo:
            self.caminho_planilha = arquivo
            self.lbl_caminho.config(text=os.path.basename(arquivo), foreground="black", font=("Segoe UI", 10, "bold"))
            self.atualizar_status("Planilha vinculada com sucesso.", INFO)

    def limpar_status_planilha(self):
        if not self.caminho_planilha:
            messagebox.showwarning("Aviso", "Por favor, selecione a planilha de pendências primeiro.")
            return

        resposta = messagebox.askyesno("Limpar Status", f"Tem certeza que deseja apagar todos os registros 'Enviado' e 'Não Encontrado' do arquivo selecionado?\n\nArquivo atual: {os.path.basename(self.caminho_planilha)}")
        if not resposta:
            return

        try:
            wb = load_workbook(self.caminho_planilha)
            sh_ROBO = wb["ROBO"]
            
            coluna_status = 7
            for col in range(1, sh_ROBO.max_column + 1):
                if str(sh_ROBO.cell(row=8, column=col).value).strip() == "Status":
                    coluna_status = col
                    break
            
            linhas_apagadas = 0
            for row in range(9, sh_ROBO.max_row + 1):
                valor = str(sh_ROBO.cell(row=row, column=coluna_status).value)
                if valor in ["Enviado", "Não Encontrado"]:
                    sh_ROBO.cell(row=row, column=coluna_status).value = ""
                    linhas_apagadas += 1
                    
            wb.save(self.caminho_planilha)
            messagebox.showinfo("Sucesso", f"Status limpos com sucesso!\n{linhas_apagadas} linhas foram resetadas.")
            self.atualizar_status(f"Planilha resetada ({linhas_apagadas} linhas). Pronto para novo teste.", SUCCESS)
            
        except PermissionError:
            messagebox.showerror("Erro de Permissão", "A planilha está aberta no Excel! Feche o arquivo e tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao limpar a planilha:\n{e}")

    def iniciar_processo_login(self):
        nome = Querybox.get_string(title="Autenticação Teams", prompt="Quem vai utilizar o robô agora? (Digite seu nome ou e-mail):", initialvalue="")
        if nome: threading.Thread(target=robo_login, args=(self, nome), daemon=True).start()

    def iniciar_disparos(self):
        if not self.caminho_planilha:
            messagebox.showwarning("Aviso", "Por favor, selecione a planilha de pendências antes de iniciar o envio.")
            return
            
        diretorio = os.path.dirname(self.caminho_planilha)
        nome_arquivo = os.path.basename(self.caminho_planilha)
        
        if not nome_arquivo.endswith("_Copia_Segura.xlsx"):
            nome_sem_ext, ext = os.path.splitext(nome_arquivo)
            caminho_copia = os.path.join(diretorio, f"{nome_sem_ext}_Copia_Segura{ext}")
            
            try:
                shutil.copy2(self.caminho_planilha, caminho_copia)
                self.caminho_planilha = caminho_copia
                self.lbl_caminho.config(text=os.path.basename(caminho_copia), foreground="blue")
                self.atualizar_status(f"Cópia de segurança criada e em uso.", INFO)
            except Exception as e:
                messagebox.showerror("Erro de Cópia", f"Não foi possível criar a cópia de segurança:\n{e}")
                return

        texto_atual = self.txt_mensagem.get("1.0", tk.END).strip()
        salvar_template(self.tipo_ativo, texto_atual) 
        
        if self.tipo_ativo == "sedex" and "[TABELA_RASTREIO]" not in texto_atual:
            if not messagebox.askyesno("Cuidado!", "Você apagou a tag [TABELA_RASTREIO] do Sedex. A tabela não será gerada.\nEnviar mesmo assim?"): return
        elif "{primeiro_nome}" not in texto_atual:
            if not messagebox.askyesno("Cuidado!", "Você apagou a tag {primeiro_nome}. A mensagem não será personalizada com o nome do colaborador.\nEnviar mesmo assim?"): return

        threading.Thread(target=robo_disparos, args=(self, self.caminho_planilha, texto_atual), daemon=True).start()

if __name__ == "__main__":
    app = AppTeams()
    app.mainloop()